# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
import collections
import re

alignment_center=Alignment(horizontal='center', vertical='bottom', text_rotation=0,wrap_text=False,shrink_to_fit=False,indent=0)

def worker(sourceFile,targetFile):
    wb = load_workbook(sourceFile)
    ws = wb.active
    # fetch given excel file, save all usefull infor into all_dict
    all_dic = parseSourceFile(ws)
    if not len(all_dic):
        return False
    # ws1: parse given input movie information(all_dic), insert into targetFile
    try:
        target_wb = load_workbook(targetFile)
        if ws.title in target_wb.get_sheet_names():
            target_wb.remove_sheet(target_wb[ws.title])
        ws1 = target_wb.create_sheet()
        ws1.title = ws.title
    except Exception, e:
        print ("Throw exception: %s" % str(e))
        print "Target file not existed, start create a new excel."
        target_wb = Workbook()
        ws1 = target_wb.active
        ws1.title = ws.title
    buildMoiveSheet(ws1,all_dic)

    # ws2: creat or update "报名" sheet ############################
    if u'报名' in [sheet.title for sheet in target_wb.worksheets]:
        ws2 = target_wb.get_sheet_by_name(u'报名')
    else:
        ws2 = target_wb.create_sheet()
        ws2.title = u'报名'

    #get active position for insert new data
    for i in ws2[1]:
        if i.value == ws1.title:
            pos = ws2[1].index(i) + 1
            statistic_pos = len(ws2[1]) - 1
            evaluate_pos = len(ws2[1])
            break
    else:
        if len(ws2[1]) == 1:
            pos = 1
        else:
            pos = len(ws2[1]) - 2 + 1
        statistic_pos = pos + 1
        evaluate_pos = pos + 2
    city_list = ['安阳','包头','保定','北海','北京','常州','郴州','成都','赤峰','大连','大庆','大同','丹东','东莞','东营','佛山','福清','福州','抚顺','广元','贵阳','桂林','哈尔滨','汉中','杭州','合肥','衡阳','呼和浩特','淮北','淮南','鸡西','吉林','济南','济宁','佳木斯','江门','江阴','金华','晋城','晋江','荆门','荆州','昆明','兰州','临汾','临淄','柳州','龙岩','洛阳','马鞍山','绵阳','牡丹江','南京','南宁','南通','宁波','宁德','盘锦','莆田','齐齐哈尔','青岛','青岛骑士30','泉州','厦门','绍兴','沈阳','石家庄','寿光','四木','台州','太原','泰州','唐山','天津','天津骑士30','铜陵','万州','威海','潍坊','温州','无锡','芜湖','武汉','西安','湘潭','徐州','烟台','延吉','宜昌','宜兴','银川','营口','湛江','漳州','长春','长沙','镇江','重庆','株洲','淄博']

    # clear old columns data
    clear_columns(ws2, pos)
    # insert or update new item
    ws2.cell(row=1, column=pos).value = ws1.title
    shift = 2
    for i in all_dic.iterkeys():
        ws2.cell(row=shift, column=pos).value = i
        shift = shift + 1
    # update data for 统计，满报值
    ws2.cell(row=1, column=statistic_pos).value = '统计'
    ws2.cell(row=1, column=evaluate_pos).value = '满报值'
    statistic_dict = {} #for ws3 use(影响力统计->满报奖励)
    shift = 2
    for i in city_list:
        ws2.cell(row=shift, column=statistic_pos).value = i
        count = 0
        for ro in range(2, len(city_list) + 2):
            for col in range(1, statistic_pos):
                if ws2.cell(row=ro, column=col).value == i.decode('utf-8'):
                    count += 1
        if count >= statistic_pos - 1:
            ws2.cell(row=shift, column=evaluate_pos).value = 50
        else:
            ws2.cell(row=shift, column=evaluate_pos).value = 0
        statistic_dict[i.decode('utf-8')] = ws2.cell(row=shift, column=evaluate_pos).value
        shift += 1

    #ws3: update last sheet "影响力统计" ############################
    if u'影响力统计' in [sheet.title for sheet in target_wb.worksheets]:
        target_wb.remove_sheet(target_wb[u'影响力统计'])

    ws3 = target_wb.create_sheet()
    ws3.title = u'影响力统计'

    assign_count = statistic_pos - 1 + 1 #add 满报奖励
    total_count = assign_count + 3
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_count)
    ws3['A1'] = u'影迷团影力值统计'
    ws3['A1'].alignment = alignment_center
    ws3.merge_cells('A2:A3')
    ws3['A2'] = u'城市'
    ws3['A2'].alignment = alignment_center
    ws3.merge_cells(start_row=2, start_column=2, end_row=2, end_column=1+assign_count)
    ws3['B2'] = u'院线活动'
    ws3['B2'].alignment = alignment_center
    ws3.merge_cells(start_row=2, start_column=2+assign_count, end_row=3, end_column=2+assign_count)
    ws3.cell(row=2, column=2+assign_count).value = u'自由活动'
    ws3.merge_cells(start_row=2, start_column=total_count, end_row=3, end_column=total_count)
    ws3.cell(row=2, column=total_count).value = u'合计'

    # insert 城市 column
    shift = 4
    for i in city_list:
        ws3.cell(row=shift, column=1).value = i
        shift +=1
    # get dict data from all movies, to insert into this sheet
    statistic_data = {}
    for i in [cell.value for cell in ws2[1] if cell.value not in [u'统计',u'满报值']]:
        ws_temp = target_wb.get_sheet_by_name(i)
        item = collections.Counter()
        for j in range(3, ws_temp.max_row + 1):
            item[ws_temp['B' + str(j)].value] = ws_temp['N' + str(j)].value
        statistic_data[i] = item

    shift_col = 2
    for i in statistic_data.keys():
        ws3.cell(row=3, column=shift_col).value = i
        for j in range(4, len(city_list) + 4):
            ws3.cell(row=j, column=shift_col).value = statistic_data[i][ws3['A' + str(j)].value]
        shift_col += 1

    # get 满报奖励 from 报名 sheet
    ws3.cell(row=3, column=shift_col).value = u'满报奖励'
    for i in range(4, len(city_list) + 4):
        ws3.cell(row=i, column=shift_col).value = statistic_dict[ws3['A' + str(i)].value]

    try:
        target_wb.save(targetFile)
    except IOError:
        print(u"生成文件失败！请保证先关闭要保存的文件后，再操作->" + targetFile)
        return "File save failed"
    print("Work done, save as: " + targetFile)
    return True

def clear_columns(sheet, pos):
    for i in sheet.rows:
        i[pos-1].value = None

def locateKeyword(sheet, keyword, count=1):
    result = []
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if isinstance(sheet.cell(row=i, column=j).value, unicode) and re.search(keyword,sheet.cell(row=i, column=j).value) or sheet.cell(row=i, column=j).value == keyword:
                if count == 1:
                    return i,j
                else:
                    result.append((i,j))
    return result

def parseSourceFile(sheet):
    city_column = locateKeyword(sheet,u'城市名')
    title_column = locateKeyword(sheet, u'结案人职务')
    people_column = locateKeyword(sheet, u'合影人数')
    ma_column = locateKeyword(sheet, u'码数')
    ticket_column = locateKeyword(sheet, u'票根数量')
    score_column = locateKeyword(sheet, u'影评总计')
    link_column = locateKeyword(sheet, u'结案链接')

    all_dic = {}
    if not min(len(city_column), len(title_column), len(people_column), len(ma_column), len(ticket_column), len(score_column), len(link_column)):
        print u"""缺少必要信息，解析失败！！
请检查文件，确保"城市名", "结案人职务","合影人数","码数","票根数量","影评总计","结案链接" 全部存在后再操作！"""
        return all_dic

    for i in range(2, sheet.max_row + 1):
        if not sheet.cell(row=i, column=city_column[1]).value:
            continue
        item = {}
        item['title'] = sheet.cell(row=i, column=title_column[1]).value
        item['people_num'] = sheet.cell(row=i, column=people_column[1]).value
        item['ma_num'] = sheet.cell(row=i, column=ma_column[1]).value
        item['ticket_num'] = sheet.cell(row=i, column=ticket_column[1]).value
        item['score'] = sheet.cell(row=i, column=score_column[1]).value
        item['app_link'] = sheet.cell(row=i, column=link_column[1]).value
        all_dic[sheet.cell(row=i, column=city_column[1]).value] = item
    return all_dic

def buildMoiveSheet(sheet,all_dic):
    sheet.cell(row=1, column=1).value = sheet.title
    sheet.cell(row=1, column=2).value = '误差率:'
    sheet.cell(row=2, column=1).value = '编号'
    column_list = ['编号', '城市名', '结案人职务', '码数', '合影', '误差', '票根', '误差', '影评', '误差', '结果', '影力值1', '影力值2', '影力值3', '结案链接']
    # build column
    for i in range(1, len(column_list) + 1):
        sheet.cell(row=2, column=i).value = column_list[i - 1]

    shift = 3
    num = 1  # 编号
    for city in all_dic.iterkeys():
        sheet.cell(row=shift, column=1).value = num
        sheet.cell(row=shift, column=2).value = city
        sheet.cell(row=shift, column=3).value = all_dic[city]['title']
        sheet.cell(row=shift, column=4).value = all_dic[city]['ma_num']
        sheet.cell(row=shift, column=5).value = all_dic[city]['people_num']
        column6 = int(round(float(all_dic[city]['people_num']) / all_dic[city]['ma_num'] * 100))
        sheet.cell(row=shift, column=6).value = str(column6) + '%'
        sheet.cell(row=shift, column=7).value = all_dic[city]['ticket_num']
        column8 = int(round(float(all_dic[city]['ticket_num']) / all_dic[city]['ma_num'] * 100))
        sheet.cell(row=shift, column=8).value = str(column8) + '%'
        sheet.cell(row=shift, column=9).value = all_dic[city]['score']
        column10 = int(round(float(all_dic[city]['score']) / all_dic[city]['ma_num'] * 100))
        sheet.cell(row=shift, column=10).value = str(column10) + '%'
        if column6 < 85 or column8 < 85 or column10 < 85:
            sheet.cell(row=shift, column=11).value = '不合格'
            sheet.cell(row=shift, column=12).value = 0
            sheet.cell(row=shift, column=13).value = 0
            sheet.cell(row=shift, column=14).value = 0
        else:
            sheet.cell(row=shift, column=11).value = '合格'
            sheet.cell(row=shift, column=12).value = 100
            sheet.cell(row=shift, column=13).value = 0
            sheet.cell(row=shift, column=14).value = 100
            if all_dic[city]['title'] == u'影迷团长':
                sheet.cell(row=shift, column=13).value = 10
                sheet.cell(row=shift, column=14).value = 110
        sheet.cell(row=shift, column=15).value = all_dic[city]['app_link']
        shift = shift + 1
        num = num + 1
