# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'untitled_main.ui'
#
# Created by: PyQt4 UI code generator 4.11.4
#
# WARNING! All changes made in this file will be lost!
from PyQt4 import QtCore, QtGui
from PyQt4.QtGui import *
import winsound
from openpyxl import load_workbook
import re

try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

try:
    _encoding = QtGui.QApplication.UnicodeUTF8
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig, _encoding)
except AttributeError:
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig)

class Ui_MainWindow(QtGui.QMainWindow):
    def __init__(self):
        super(Ui_MainWindow, self).__init__()
        self.setupUi(self)
        self.retranslateUi(self)

    def setupUi(self, MainWindow):
        MainWindow.setObjectName(_fromUtf8("MainWindow"))
        MainWindow.resize(1120, 838)
        self.centralwidget = QtGui.QWidget(MainWindow)
        self.centralwidget.setObjectName(_fromUtf8("centralwidget"))
        self.plainTextEdit = QtGui.QPlainTextEdit(self.centralwidget)
        self.plainTextEdit.setGeometry(QtCore.QRect(20, 60, 1081, 761))
        self.plainTextEdit.setObjectName(_fromUtf8("plainTextEdit"))
        self.verticalScrollBar = QtGui.QScrollBar(self.centralwidget)
        self.verticalScrollBar.setStyleSheet(_fromUtf8("    background: #8000FF;\n"
                                                       "    border: 3px solid grey;\n"
                                                       "    border-radius:5px;\n"
                                                       "    min-height: 20px;"))
        self.plainTextEdit.setVerticalScrollBar(self.verticalScrollBar)
        #self.pushButton = QtGui.QPushButton(self.centralwidget)
        #self.pushButton.setGeometry(QtCore.QRect(850, 20, 112, 34))
        #self.pushButton.setObjectName(_fromUtf8("pushButton"))
        #self.pushButton.clicked.connect(self.browse)
        #self.comboBox = QtGui.QComboBox(self.centralwidget)
        #self.comboBox.setGeometry(QtCore.QRect(20, 20, 761, 27))
        #self.comboBox.setObjectName(_fromUtf8("comboBox"))
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtGui.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1120, 31))
        self.menubar.setObjectName(_fromUtf8("menubar"))
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtGui.QStatusBar(MainWindow)
        self.statusbar.setObjectName(_fromUtf8("statusbar"))
        MainWindow.setStatusBar(self.statusbar)

        #redirect console to plainText
        redir = redirect(self.plainTextEdit)
        sys.stdout = redir
        sys.stderr = redir

        # Enable dragging and dropping onto the GUI
        self.setAcceptDrops(True)
        # Enable Copy&Paste
        self.setupEditActions()
        self.actionCut.setEnabled(False)
        self.actionCopy.setEnabled(False)
        self.actionCut.triggered.connect(self.plainTextEdit.cut)
        self.actionCopy.triggered.connect(self.plainTextEdit.copy)
        self.actionPaste.triggered.connect(self.plainTextEdit.paste)
        self.plainTextEdit.copyAvailable.connect(self.actionCut.setEnabled)
        self.plainTextEdit.copyAvailable.connect(self.actionCopy.setEnabled)
        QtGui.QApplication.clipboard().dataChanged.connect(self.clipboardDataChanged)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        MainWindow.setWindowTitle(_translate("MainWindow", "Excel_Tool", None))
        #self.pushButton.setText(_translate("MainWindow", "Brower", None))
        self.lineEdit.setText(_translate("Dialog", "统计值", None))
        self.label.setText(_translate("Dialog", "TargetFile", None))

    def browse(self):
        #fileName = QtGui.QFileDialog.getOpenFileName(self, "Open File", '.',"HTML Files (*.htm *.html)")
        fileName = QtGui.QFileDialog.getOpenFileName(self, "Open File", '.', "EXCEL Files (*.xlsx)")
        if fileName:
            if self.comboBox.findText(fileName) == -1:
                self.comboBox.addItem(fileName)
            self.comboBox.setCurrentIndex(self.comboBox.findText(fileName))
            self.start()

    def start(self):
        if self.comboBox.currentText() == '':
            print("Please brower a valid html")
        else:
            self.plainTextEdit.moveCursor(QtGui.QTextCursor.End)
            self.th = workThread(self.comboBox.currentText())
            self.th.started.connect(self.th.commentCheck)
            self.th.start()

    # The following three methods set up dragging and dropping for the app
    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            e.accept()
        else:
            e.ignore()

    def dropEvent(self, e):
        if e.mimeData().hasUrls():
            e.accept()
            self.plainTextEdit.moveCursor(QtGui.QTextCursor.End)
            for url in e.mimeData().urls():
                #fname = str(url.toLocalFile())
                fname = unicode(url.toLocalFile())
                print("Input file: " + fname)
                self.th = workThread(fname)
                self.th.started.connect(self.th.commentCheck)
                self.th.start()
                break
        else:
            e.ignore()

    def setupEditActions(self):
        self.actionCopy = QtGui.QAction(
                "&Copy", self, priority=QtGui.QAction.LowPriority,
                shortcut=QtGui.QKeySequence.Copy)

        self.actionCut = QtGui.QAction(
                "&Cut", self, priority=QtGui.QAction.LowPriority,
                shortcut=QtGui.QKeySequence.Cut)

        self.actionPaste = QtGui.QAction(
                "&Paste", self, priority=QtGui.QAction.LowPriority,
                shortcut=QtGui.QKeySequence.Paste,
                enabled=(len(QtGui.QApplication.clipboard().text()) != 0))

    def clipboardDataChanged(self):
        self.actionPaste.setEnabled(len(QtGui.QApplication.clipboard().text()) != 0)

class redirect(object):
    def __init__(self, text):
        self.output = text

    def write(self, string, color = False):
        fmt = QTextCharFormat()
        if color:
            fmt.setForeground(QColor("red"))
        else:
            fmt.setForeground(QColor("black"))

        self.output.mergeCurrentCharFormat(fmt)
        self.output.insertPlainText(string)
        self.output.ensureCursorVisible()

class workThread(QtCore.QThread):
    def __init__(self, file = None):
        super(workThread, self).__init__()
        self.file = file

    def commentCheck(self):
        winsound.PlaySound('alert', winsound.SND_ASYNC)
        worker(self.file)
        self.exit()

def worker(file):
    wb = load_workbook(file)
    for sheet in wb.get_sheet_names():
        ws = wb.get_sheet_by_name(sheet)

        for i in range(1, ws.max_row + 1):
            for j in range(1, ws.max_column + 1):
                string = ws.cell(row=i, column=j).value
                if isinstance(string,long) and re.match(r'1[3458]\d{9}', str(string)):
                    ws.cell(row=i, column=j).value = re.sub(
                        r'(1[3458][0-9])(\d{4})(\d{4})', r'\1****\3', str(string))
                    print ws.cell(row=i, column=j).value
    wb.save(file)
    print("Work done, save as: " + file)

if __name__ == "__main__":
    import sys
    app = QtGui.QApplication(sys.argv)
    splash = QSplashScreen(QPixmap("spinner.gif"))
    splash.show()
    splash.showMessage("loading...")
    app.processEvents() #in case GUI not response, enable processEvent
    ui = Ui_MainWindow()
    ui.show()
    splash.finish(ui)
    sys.exit(app.exec_())

